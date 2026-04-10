import os
import yaml
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from shutil import copyfile

# ─────────────────────────────────────────────────────────────────────────────
# Config
# ─────────────────────────────────────────────────────────────────────────────

def load_config(config_path: str) -> dict:
    """Load a YAML configuration file and return as dict."""
    if not os.path.exists(config_path):
        raise FileNotFoundError(f"Config file not found: {config_path}")
    with open(config_path, "r", encoding="utf-8") as f:
        config = yaml.safe_load(f)
    if config is None:
        raise ValueError(f"Config file is empty or invalid YAML: {config_path}")
    return config


def resolve_path(base_dir: str, path: str) -> str:
    """Resolve path relative to base_dir if not already absolute."""
    if os.path.isabs(path):
        return path
    return os.path.abspath(os.path.join(base_dir, path))


# ─────────────────────────────────────────────────────────────────────────────
# Template Reading
# ─────────────────────────────────────────────────────────────────────────────

def _deduplicate_headers(headers: list) -> list:
    """
    Deduplicate column names by appending .1, .2, ... to duplicates.
    Mirrors pandas behaviour so the cleaner can strip them later.
    """
    seen = {}
    result = []
    for h in headers:
        if h in seen:
            seen[h] += 1
            result.append(f"{h}.{seen[h]}")
        else:
            seen[h] = 0
            result.append(h)
    return result


def read_template_sheet(file_path: str, sheet_name: str) -> tuple:
    """
    Read one data sheet from the template Excel file.

    Template structure:
        Row 1  — friendly descriptions  (ignored as header, kept as label map)
        Row 2  — technical field names  (used as column headers)
        Row 3+ — actual data rows

    Returns:
        df        : DataFrame with row-2 headers; data starts at index 0.
                    Includes a hidden '_row_idx' column (0-based) for
                    writing results back to the correct Excel row.
        label_map : dict {technical_name: friendly_description}
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Template not found: {file_path}")

    wb = load_workbook(file_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Sheet '{sheet_name}' not found in '{file_path}'")

    ws = wb[sheet_name]
    all_rows = [
        [cell if cell is not None else "" for cell in row]
        for row in ws.iter_rows(values_only=True)
    ]
    wb.close()

    if len(all_rows) < 2:
        raise ValueError(
            f"Sheet '{sheet_name}' must have at least 2 rows "
            f"(row 1 = descriptions, row 2 = headers)."
        )

    desc_row    = [str(v).strip() for v in all_rows[0]]
    header_row  = [str(v).strip() if str(v).strip() else f"col_{i}"
                   for i, v in enumerate(all_rows[1])]
    data_rows   = all_rows[2:]

    header_row = _deduplicate_headers(header_row)

    # friendly label: technical_name -> description
    label_map = {h: d for h, d in zip(header_row, desc_row) if h}

    df = pd.DataFrame(data_rows, columns=header_row)
    df = df.fillna("").astype(str).replace("None", "")

    # 0-based index → Excel row = _row_idx + 3  (header at row 2, data at row 3+)
    df["_row_idx"] = range(len(df))

    return df, label_map


def get_sheet_names(file_path: str, exclude: list = None) -> list:
    """Return all sheet names from an Excel file, excluding those in `exclude`."""
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")
    exclude = [s.lower() for s in (exclude or [])]
    wb = load_workbook(file_path, read_only=True)
    names = [s for s in wb.sheetnames if s.lower() not in exclude]
    wb.close()
    return names

def get_ignore_fields_from_template(file_path: str, ignore_sheet_name: str) -> dict:
    """
    Read the ignore-fields configuration sheet from the template.

    Expected layout (column-based, one column per JOB):
        Row 1  — column headers: "GLOBAL" | "<JOB NAME 1>" | "<JOB NAME 2>" | ...
        Row 2+ — field names to ignore for that column's scope

    - "GLOBAL" column  : fields ignored for every JOB
    - JOB-name columns : fields ignored only for that specific JOB

    Example sheet:
        | GLOBAL   | Maintenance Order Header | Maintenance Operation |
        | FIELD_A  | FIELD_X                  | FIELD_Y               |
        | FIELD_B  |                          | FIELD_Z               |

    Returns:
        dict mapping scope → list[str], e.g.
        {
            "GLOBAL":                     ["FIELD_A", "FIELD_B"],
            "Maintenance Order Header":   ["FIELD_X"],
            "Maintenance Operation":      ["FIELD_Y", "FIELD_Z"],
        }

    Use get_effective_ignore_fields() to resolve the merged list for a specific JOB.
    """
    wb = load_workbook(file_path, read_only=True, data_only=True)
    if ignore_sheet_name not in wb.sheetnames:
        wb.close()
        print(f"  ⚠️  Ignore-field sheet '{ignore_sheet_name}' not found — no fields will be ignored from template.")
        return {}

    ws = wb[ignore_sheet_name]
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not all_rows:
        return {}

    # Row 1 = column headers (scope names)
    headers = [str(cell).strip() if cell is not None else "" for cell in all_rows[0]]

    result: dict = {}
    for col_idx, scope in enumerate(headers):
        if not scope:
            continue
        fields = [
            str(row[col_idx]).strip()
            for row in all_rows[1:]
            if col_idx < len(row)
            and row[col_idx] is not None
            and str(row[col_idx]).strip()
        ]
        result[scope] = fields

    return result


def get_effective_ignore_fields(ignore_fields_map: dict, job_name: str) -> list:
    """
    Merge GLOBAL ignore fields with job-specific ignore fields.

    Args:
        ignore_fields_map : dict returned by get_ignore_fields_from_template()
        job_name          : exact JOB NAME string from config (e.g. "Maintenance Operation")

    Returns:
        Deduplicated list of field names to ignore for this JOB.
    """
    global_fields   = ignore_fields_map.get("GLOBAL", [])
    job_fields      = ignore_fields_map.get(job_name, [])
    return list(dict.fromkeys(global_fields + job_fields))  # preserve order, deduplicate

# ─────────────────────────────────────────────────────────────────────────────
# Data Standard
# ─────────────────────────────────────────────────────────────────────────────

def load_data_standard(ds_path: str, ds_sheet: str) -> pd.DataFrame:
    """Load the Data Standard Excel file and return as a DataFrame."""
    if not os.path.exists(ds_path):
        raise FileNotFoundError(f"Data standard file not found: {ds_path}")
    df = pd.read_excel(ds_path, sheet_name=ds_sheet, header=0, dtype=str, keep_default_na=False)
    df.columns = [str(c).strip() for c in df.columns]
    return df


def _find_ds_columns(ds_df: pd.DataFrame) -> dict:
    """
    Locate key columns in the Data Standard by keyword matching.

    Actual columns in DataStandardList.xlsx (Work Order sheet):
        'S4 Target Database Table Name'   → table  (for per-JOB filtering)
        'S4 Target Database Field Name'   → field
        'S4 Target Data Length'           → length
        'Field Requirement'               → req
        NOTE: No 'Data Type' column exists → type is always derived as None
    """
    cols = list(ds_df.columns)
    return {
        "table":  next((c for c in cols if "Table Name" in c and "Target" in c), None),
        "field":  next((c for c in cols if "Field Name" in c and "Target" in c), None)
                  or next((c for c in cols if "Field Name" in c), None),
        "length": next((c for c in cols if "Data Length" in c), None),
        "req":    next((c for c in cols if "Field Requirement" in c or "Requirement" in c), None),
    }


def get_field_metadata(
    ds_df: pd.DataFrame,
    field_name: str,
    ds_col_map: dict,
    ds_table=None,
) -> dict:
    """
    Look up one field's validation metadata from the Data Standard.

    Strips known column suffixes (_TGT, _SRC, _DESC) to get the base SAP field name,
    then searches the Data Standard filtered by `ds_table` (if provided).

    Args:
        ds_df      : Full Data Standard DataFrame
        field_name : Template column name (e.g. 'AUFNR_TGT')
        ds_col_map : Result of _find_ds_columns()
        ds_table   : SAP table name(s) to filter rows. Accepts:
                     - str  : single table, e.g. 'AFIH'
                     - list : multiple tables, e.g. ['IFLOT', 'ILOA']
                     - None : search all rows (no filter)

    Returns:
        dict with keys: mandatory (bool), length (int|None), type (str|None)
    """
    # Strip known template suffixes to get SAP base field name
    base = field_name
    for suffix in ("_TGT", "_SRC", "_DESC", "_tgt", "_src", "_desc"):
        if base.endswith(suffix):
            base = base[: -len(suffix)]
            break

    field_col  = ds_col_map.get("field")
    table_col  = ds_col_map.get("table")
    length_col = ds_col_map.get("length")
    req_col    = ds_col_map.get("req")

    if not field_col:
        return {"mandatory": False, "length": None, "type": None}

    # Filter by table name(s) — supports str or list
    df_scope = ds_df
    if ds_table and table_col:
        if isinstance(ds_table, list):
            tables_upper = [t.strip().upper() for t in ds_table]
            df_scope = ds_df[ds_df[table_col].str.strip().str.upper().isin(tables_upper)]
        else:
            df_scope = ds_df[ds_df[table_col].str.strip().str.upper() == ds_table.strip().upper()]

    mask = df_scope[field_col].str.strip().str.upper() == base.upper()
    rows = df_scope[mask]
    if rows.empty:
        return {"mandatory": False, "length": None, "type": None}

    row = rows.iloc[0]

    # Mandatory: starts with "M-Mandatory" or just "M"
    req_str = str(row.get(req_col, "") if req_col else "").strip().upper()
    is_mandatory = req_str.startswith("M-MANDATORY") or req_str == "M"

    # Length
    length = None
    if length_col:
        try:
            length = int(float(str(row[length_col]).strip()))
        except (ValueError, TypeError):
            pass

    # NOTE: No 'Data Type' column in DataStandardList.xlsx — always None
    return {"mandatory": is_mandatory, "length": length, "type": None}


def build_field_metadata(
    ds_df: pd.DataFrame,
    data_columns: list,
    ds_table=None,
) -> dict:
    """
    Build {col_name: metadata_dict} for every column in `data_columns`.

    Args:
        ds_table: SAP table name(s) to filter the Data Standard.
                  Accepts str (e.g. 'AFIH') or list (e.g. ['IFLOT','ILOA']).
                  Comes from DS_TABLE in each JOB config.
    """
    ds_col_map = _find_ds_columns(ds_df)
    return {
        col: get_field_metadata(ds_df, col, ds_col_map, ds_table=ds_table)
        for col in data_columns
    }
    
# ─────────────────────────────────────────────────────────────────────────────
# KDS Mapping
# ─────────────────────────────────────────────────────────────────────────────

def load_kds_mapping(mapping_path: str) -> dict:
    """Load all sheets from the KDS Mapping Excel file. Returns {sheet_name: DataFrame}."""
    if not os.path.exists(mapping_path):
        raise FileNotFoundError(f"KDS mapping file not found: {mapping_path}")
    sheets = pd.read_excel(mapping_path, sheet_name=None, dtype=str, keep_default_na=False)
    return {name: df for name, df in sheets.items()}


# ─────────────────────────────────────────────────────────────────────────────
# Column Filtering
# ─────────────────────────────────────────────────────────────────────────────

def should_ignore_column(col_name: str, ignore_fields: list, ignore_suffixes: list) -> bool:
    """Return True if a column should be excluded from validation."""
    col_upper = col_name.upper()
    # Exact match (case-insensitive)
    if col_upper in {f.upper() for f in ignore_fields}:
        return True
    # Suffix match
    for suffix in ignore_suffixes:
        if col_upper.endswith(suffix.upper()):
            return True
    return False


def get_data_columns(df: pd.DataFrame, ignore_fields: list, ignore_suffixes: list) -> list:
    """Return the ordered list of columns that should be validated."""
    return [
        col for col in df.columns
        if col != "_row_idx"
        and not col.endswith(".1")          # pandas duplicate suffix
        and not should_ignore_column(col, ignore_fields, ignore_suffixes)
    ]




# ─────────────────────────────────────────────────────────────────────────────
# Output / Writing
# ─────────────────────────────────────────────────────────────────────────────

def save_results_to_excel(
    template_path: str,
    df_results: dict,
    output_path: str,
    header_excel_row: int = 2,
    data_excel_start_row: int = 3,
) -> str:
    """
    Write validation check columns to a copy of the template Excel file.

    Args:
        template_path        : Source template (.xlsx) to copy from
        df_results           : {sheet_name: DataFrame} — each df has Check* columns
                               and a '_row_idx' column (0-based)
        output_path          : Where to save the annotated Excel file
        header_excel_row     : Excel row (1-indexed) where column headers live (default 2)
        data_excel_start_row : Excel row (1-indexed) where data rows start   (default 3)

    Returns:
        Absolute path to the saved file
    """
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    copyfile(template_path, output_path)

    wb = load_workbook(output_path)

    header_font  = Font(bold=True)
    header_fill  = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    thin_border  = Border(
        left   = Side(style="thin", color="808080"),
        right  = Side(style="thin", color="808080"),
        top    = Side(style="thin", color="808080"),
        bottom = Side(style="thin", color="808080"),
    )
    wrap_align = Alignment(wrapText=True)

    for sheet_name, df in df_results.items():
        if sheet_name not in wb.sheetnames:
            print(f"  ⚠️  Sheet '{sheet_name}' not found in workbook — skipping.")
            continue

        check_cols = [c for c in df.columns if c.startswith("Check ")]
        if not check_cols:
            continue

        ws = wb[sheet_name]
        start_col_idx = ws.max_column + 1

        for offset, col_name in enumerate(check_cols):
            col_idx = start_col_idx + offset

            # Header cell (row 2 = technical field names row)
            hdr_cell = ws.cell(row=header_excel_row, column=col_idx, value=col_name)
            hdr_cell.font      = header_font
            hdr_cell.fill      = header_fill
            hdr_cell.border    = thin_border
            hdr_cell.alignment = wrap_align

            # Data cells
            for _, row in df.iterrows():
                try:
                    row_idx = int(row["_row_idx"])
                except (KeyError, TypeError, ValueError):
                    continue

                excel_row = row_idx + data_excel_start_row   # 0-based → 1-indexed Excel
                value     = row.get(col_name)
                if pd.isna(value) if isinstance(value, float) else (value is None):
                    value = None

                data_cell = ws.cell(row=excel_row, column=col_idx, value=value)
                data_cell.border    = thin_border
                data_cell.alignment = wrap_align

            # Width = header text length + padding
            col_letter = hdr_cell.column_letter
            ws.column_dimensions[col_letter].width = len(col_name) + 2

    wb.save(output_path)
    wb.close()
    return os.path.abspath(output_path)